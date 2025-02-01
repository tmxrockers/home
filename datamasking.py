# # # import pandas as pd

# # # file_path = r"\\192.168.0.114\DataMasking\GDCE Resource Skill Details.xlsx"
# # # df = pd.read_excel(file_path)
# # # print(df.head())


# import smbclient
# import pandas as pd
# import io

# # SMB Authentication
# smbclient.ClientConfig(username="Arun", password="9655299619")

# # File path on the shared drive
# smb_path = r"\\192.168.0.114\DataMasking\GDCE Resource Skill Details.xlsx"

# # Open file directly
# with smbclient.open_file(smb_path, mode="rb") as f:
#     df = pd.read_excel(io.BytesIO(f.read()))

# print(df.head())


# import io
# import os
# import openpyxl
# from smbprotocol.connection import Connection
# from smbprotocol.session import Session
# from smbprotocol.tree import TreeConnect
# from smbprotocol.open import Open, FilePipePrinterAccessMask, CreateDisposition

# # SMB Configuration
# server = "192.168.0.114"  # Example: "192.168.1.100"
# share = "DataMasking"   # Example: "SharedFolder"
# username = "Arun"
# password = "9655299619"
# file_path = "GDCE Resource Skill Details.xlsx"  # Path inside the shared folder

# # Function to apply masking patterns
# def apply_masking(value, pattern, mask_value):
#     """Apply different masking patterns."""
#     if value is None or str(value).strip() == "":
#         return value  

#     if pattern == "Constant":
#         return mask_value
#     elif pattern == "Prefix":
#         return f"{mask_value}{value}"
#     elif pattern == "Suffix":
#         return f"{value}{mask_value}"
#     elif pattern == "Email":
#         return mask_email(str(value))
#     elif pattern == "Null":
#         return None
#     return value  

# def mask_email(value):
#     """Mask an email address while keeping the domain intact."""
#     if "@" in value:
#         username, domain = value.split("@", 1)
#         masked_username = "".join("x" if c.isalpha() else c for c in username)
#         return f"{masked_username}@{domain}"
#     return value  

# def ensure_text_format(sheet, header_row):
#     """Ensure all data is treated as text."""
#     for row in sheet.iter_rows(min_row=header_row + 1):
#         for cell in row:
#             if isinstance(cell.value, (int, float)):
#                 continue
#             if isinstance(cell.value, str) and any(fmt in cell.number_format for fmt in ["MM/dd/yyyy", "yyyy-mm-dd", "d-mmm-yy"]):
#                 cell.value = str(cell.value)  
#                 cell.number_format = "@"
#             elif isinstance(cell.value, str):
#                 cell.value = cell.value.strip()
#             cell.number_format = "@"

# def ensure_valid_headers(sheet, header_row):
#     """Ensure headers are valid and string-based."""
#     headers = []
#     for cell in sheet[header_row]:
#         if isinstance(cell.value, (int, float)):  
#             cell.value = f"Column_{cell.column}"
#         elif cell.value is None or str(cell.value).strip() == "":
#             cell.value = f"Column_{cell.column}"
#         else:
#             cell.value = str(cell.value).strip()
#         cell.number_format = "@"
#         headers.append(cell.value)
#     return headers

# def mask_excel_file(server, share, username, password, file_path, header_row, rules):
#     """Mask Excel file directly from SMB share and save it back."""
#     try:
#         # Establish SMB connection
#         conn = Connection(server=server, port=445)
#         conn.connect()

#         # Start session
#         session = Session(conn, username=username, password=password)
#         session.connect()

#         print("Connection established successfully")

#         # Connect to shared folder
#         tree = TreeConnect(session, f"\\\\{server}\\{share}")
#         tree.connect()

#         # Open file for reading
#         file_obj = Open(tree, file_path, access_mask=FilePipePrinterAccessMask.GENERIC_READ)
#         file_obj.create(CreateDisposition.FILE_OPEN)

#         # Read file content
#         file_data = file_obj.read(1024 * 1024)  
#         file_obj.close()

#         # Load workbook
#         wb = openpyxl.load_workbook(io.BytesIO(file_data))
#         sheet = wb.active  

#         # Fix headers and text format
#         headers = ensure_valid_headers(sheet, header_row)
#         ensure_text_format(sheet, header_row)

#         # Apply masking rules
#         column_indices = {col: idx + 1 for idx, col in enumerate(headers)}
#         for row in sheet.iter_rows(min_row=header_row + 1):
#             if is_row_empty(row):
#                 continue  

#             for column_name, pattern, mask_value in rules:
#                 col_idx = column_indices.get(column_name)
#                 if col_idx:
#                     cell = row[col_idx - 1]
#                     cell.value = apply_masking(cell.value, pattern, mask_value)

#         # Save masked file to SMB share
#         masked_file_data = io.BytesIO()
#         wb.save(masked_file_data)
#         masked_file_data.seek(0)

#         # Define output path
#         masked_folder = "Masked_Files"
#         masked_file_path = f"{masked_folder}/Masked_{os.path.basename(file_path)}"

#         # Create directory if not exists
#         try:
#             masked_folder_obj = Open(tree, masked_folder, access_mask=FilePipePrinterAccessMask.FILE_LIST_DIRECTORY)
#             masked_folder_obj.create(CreateDisposition.FILE_OPEN_IF)
#             masked_folder_obj.close()
#         except Exception as e:
#             print(f"Could not create/find masked folder: {e}")

#         # Write masked file
#         masked_file_obj = Open(tree, masked_file_path, access_mask=FilePipePrinterAccessMask.GENERIC_WRITE)
#         masked_file_obj.create(CreateDisposition.FILE_OVERWRITE_IF)
#         masked_file_obj.write(masked_file_data.read())
#         masked_file_obj.close()

#         print(f"Masked file saved: \\\\{server}\\{share}\\{masked_file_path}")

#         # Cleanup
#         tree.disconnect()
#         session.disconnect()
#         conn.disconnect()

#     except Exception as e:
#         print(f"Error processing {file_path}: {e}")

# # Helper function to check if a row is empty
# def is_row_empty(row):
#     """Check if all values in a row are NULL or empty."""
#     return all(cell.value is None or str(cell.value).strip() == "" for cell in row)

# # Example usage:
# masking_rules = [
#     ("Name", "Prefix", "PR-"),
#     ("Email", "Email", ""),
#     ("Phone", "Suffix", "-1234")
# ]

# mask_excel_file(server, share, username, password, file_path, header_row=2, rules=masking_rules)


# import io
# import os
# import openpyxl
# from smbprotocol.connection import Connection
# from smbprotocol.session import Session
# from smbprotocol.tree import TreeConnect
# from smbprotocol.open import Open, FilePipePrinterAccessMask, CreateDisposition

# # # SMB Configuration
# server = "192.168.0.114"  # Example: "192.168.1.100"
# share = "DataMasking"   # Example: "SharedFolder"
# username = "Arun"
# password = "9655299619"
# file_path = "GDCE Resource Skill Details.xlsx"  # Path inside the shared folder

# def mask_excel_file(server, share, username, password, file_path, header_row, rules):
#     """Mask Excel file directly from SMB share and save it back."""
#     try:
#         # Establish SMB connection
#         conn = Connection(hostname=server, port=445)
#         conn.connect()

#         # Start session
#         session = Session(conn, username=username, password=password)
#         session.connect()

#         # Connect to shared folder
#         tree = TreeConnect(session, f"\\\\{server}\\{share}")
#         tree.connect()

#         # Open file for reading
#         file_obj = Open(tree, file_path, access_mask=FilePipePrinterAccessMask.GENERIC_READ)
#         file_obj.create(CreateDisposition.FILE_OPEN)

#         # Read file content
#         file_data = file_obj.read(1024 * 1024)  # Read up to 1MB
#         file_obj.close()

#         # Load workbook
#         wb = openpyxl.load_workbook(io.BytesIO(file_data))
#         sheet = wb.active  

#         # Save masked file to SMB share
#         masked_file_data = io.BytesIO()
#         wb.save(masked_file_data)
#         masked_file_data.seek(0)

#         # Define output path
#         masked_folder = "Masked_Files"
#         masked_file_path = f"{masked_folder}/Masked_{os.path.basename(file_path)}"

#         # Write masked file
#         masked_file_obj = Open(tree, masked_file_path, access_mask=FilePipePrinterAccessMask.GENERIC_WRITE)
#         masked_file_obj.create(CreateDisposition.FILE_OVERWRITE_IF)
#         masked_file_obj.write(masked_file_data.read())
#         masked_file_obj.close()

#         print(f"Masked file saved: \\\\{server}\\{share}\\{masked_file_path}")

#         # Cleanup
#         tree.disconnect()
#         session.disconnect()
#         conn.disconnect()

#     except Exception as e:
#         print(f"Error processing {file_path}: {e}")

# # Example usage
# masking_rules = [
#     ("Name", "Prefix", "PR-"),
#     ("Email", "Email", ""),
#     ("Phone", "Suffix", "-1234")
# ]

# mask_excel_file(server, share, username, password, file_path, header_row=2, rules=masking_rules)


# import io
# import pandas as pd
# import uuid
# from smbprotocol.connection import Connection
# from smbprotocol.session import Session
# from smbprotocol.tree import TreeConnect
# from smbprotocol.open import Open, FilePipePrinterAccessMask, CreateDisposition, CreateOptions, FileAttributes, ShareAccess, ImpersonationLevel

# # SMB Authentication
# server = "192.168.0.114"
# share = "DataMasking"
# username = "Arun"
# password = "9655299619"
# file_path = "GDCE Resource Skill Details.xlsx"  # Path inside the shared folder

# # Generate a unique GUID for the connection
# connection_guid = uuid.uuid4()

# # Establish SMB Connection
# conn = Connection(guid=connection_guid, server_name=server, port=445)
# conn.connect()

# # Start Session
# session = Session(conn, username=username, password=password)
# session.connect()

# # Connect to Shared Folder
# tree = TreeConnect(session, f"\\\\{server}\\{share}")
# tree.connect()

# # Open File for Reading with required parameters
# file_obj = Open(tree, file_path)
# file_obj.create(
#     file_attributes=FileAttributes.FILE_ATTRIBUTE_NORMAL,
#     share_access=ShareAccess.FILE_SHARE_READ,
#     create_disposition=CreateDisposition.FILE_OPEN,
#     create_options=CreateOptions.FILE_NON_DIRECTORY_FILE,
#     impersonation_level=ImpersonationLevel.Impersonation,  # Set Impersonation Level
   
# )

# # Read File Content
# file_data = file_obj.read(1024 * 1024)  # Read up to 1MB
# file_obj.close()

# # Convert to DataFrame
# df = pd.read_excel(io.BytesIO(file_data))

# # Print Output
# print(df.head())

# # Cleanup
# tree.disconnect()
# session.disconnect()
# conn.disconnect()

# import os
# import pandas as pd
# from smbprotocol.connection import Connection
# from smbprotocol.session import Session
# from smbprotocol.transport import TcpTransport
# from smbprotocol.tree import Tree

# def read_excel_from_windows_share_smbprotocol(share_path, username, password):  # Removed domain
#     """Reads an Excel file from a Windows share using smbprotocol (no domain).

#     Args:
#         share_path: The full path to the Excel file.
#         username: Username for authentication.
#         password: Password for authentication.

#     Returns:
#         A pandas DataFrame, or None if an error occurs.
#     """
#     try:
#         server_name = share_path.split("\\\\")[1].split("\\")[0]
#         share_name = share_path.split("\\\\")[1].split("\\")[1]
#         file_path_in_share = "\\".join(share_path.split("\\\\")[1].split("\\")[2:])

#         transport = TcpTransport(server_name, 445)
#         connection = Connection(transport)
#         connection.connect()

#         # Session without domain
#         session = Session(connection, username=username, password=password)  # No domain here
#         session.connect()

#         tree = Tree(session, share_name)
#         tree.connect()

#         with tree.open_file(file_path_in_share, desired_access=0x80000000, share_access=1, create_disposition=5) as remote_file:
#             file_size = remote_file.get_file_size()
#             file_data = remote_file.read(file_size)

#         tree.disconnect()
#         session.disconnect()
#         connection.disconnect()
#         transport.close()

#         import tempfile
#         with tempfile.NamedTemporaryFile(delete=False) as temp_file:
#             temp_filename = temp_file.name
#             temp_file.write(file_data)

#         df = pd.read_excel(temp_filename)
#         os.remove(temp_filename)

#         return df

#     except Exception as e:
#         print(f"Error reading Excel file: {e}")
#         return None


# # Example usage (replace with your actual details)
# share_path = r"\\192.168.0.114\DataMasking\GDCE Resource Skill Details.xlsx"
# username = "Arun"
# password = "9655299619"

# df = read_excel_from_windows_share_smbprotocol(share_path, username, password)  # No domain

# if df is not None:
#     print(df.head())
# else:
#     print("Failed to read the Excel file.")

import sys
import traceback
import os
import win32wnet
import openpyxl

# Shared drive details
shared_drive_path = r"\\192.168.0.114\DataMasking"
username = "Arun"
password = "9655299619"

# Excel file path on the shared drive
excel_file_path = os.path.join(shared_drive_path, "GDCE Resource Skill Details.xlsx")

# Connect to the shared drive
try:
    win32wnet.WNetAddConnection2(
        0,  # Type of resource (0 = disk)
        None,  # Local device name (None for no redirection)
        shared_drive_path,  # Remote network name
        None,  # Provider name (None for default)
        username,
        password,
    )
    print("Connected to the shared drive successfully.")
except Exception as e:
    # Print the exception message and stack trace
    print(f"Exception occurred: {str(e)}", file=sys.stderr)
    traceback.print_exc(file=sys.stderr)
    sys.exit(1)  # Exit with a non-zero status code to indicate failure


# Read the Excel file
try:
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active

    # Example: Read data from the first row
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        print(row)

    # Example: Write data to the Excel file
    sheet.cell(row=2, column=1, value="New Data")
    workbook.save(excel_file_path)
    print("Excel file updated successfully.")

except Exception as e:
    print(f"Error reading/writing the Excel file: {e}")

# Disconnect from the shared drive
try:
    win32wnet.WNetCancelConnection2(shared_drive_path, 0, True)
    print("Disconnected from the shared drive.")
except Exception as e:
    print(f"Failed to disconnect from the shared drive: {e}")
